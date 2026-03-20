import requests
import time
import urllib.parse
import hashlib
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ================= 配置文件加载模块 =================
def load_config():
    """读取 config.json 获取 Cookie 和 User-Agent"""
    config_path = os.path.join(os.path.dirname(__file__), 'config.json')
    if not os.path.exists(config_path):
        print("[错误] 未找到 config.json！请复制 config_example.json 并填入真实的 Cookie。")
        return None
    
    with open(config_path, 'r', encoding='utf-8') as f:
        try:
            config = json.load(f)
            return config
        except json.JSONDecodeError:
            print("[错误] config.json 格式不正确，请确保是有效的 JSON 格式。")
            return None

def load_bvid_list():
    """读取需要爬取的 bvid 列表"""
    txt_path = os.path.join(os.path.dirname(__file__), 'bvid_list.txt')
    if not os.path.exists(txt_path):
        print("[错误] 未找到 bvid_list.txt！请创建并填入 BV 号。")
        return []
        
    bvid_list = []
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            # 忽略空行和注释行
            if line and not line.startswith('#'):
                bvid_list.append(line)
    return bvid_list

# ================== Wbi 签名算法加密区 ==================
mixinKeyEncTab = [
    46, 47, 18, 2, 53, 8, 23, 32, 15, 50, 10, 31, 58, 3, 45, 35, 27, 43, 5, 49,
    33, 9, 42, 19, 29, 28, 14, 39, 12, 38, 41, 13, 37, 48, 7, 16, 24, 55, 40,
    61, 26, 17, 0, 1, 60, 51, 30, 4, 22, 25, 54, 21, 56, 59, 6, 63, 57, 62, 11,
    36, 20, 34, 44, 52
]

def getMixinKey(orig: str):
    """对 img_key 和 sub_key 进行字符重新排列"""
    return ''.join([orig[i] for i in mixinKeyEncTab])[:32]

def get_wbi_keys(headers):
    """向 B站请求获取最新的 img_key 和 sub_key"""
    print("正在获取 Wbi 签名密钥...")
    url = "https://api.bilibili.com/x/web-interface/nav"
    try:
        resp = requests.get(url, headers=headers).json()
        if resp.get('code') != 0:
            print(f"获取密钥失败 (API错误): {resp.get('message')}")
            return None, None
            
        wbi_img = resp.get('data', {}).get('wbi_img', {})
        img_url = wbi_img.get('img_url')
        sub_url = wbi_img.get('sub_url')
        if not img_url or not sub_url:
            print("获取密钥失败: Wbi 数据不全")
            return None, None
            
        img_key = img_url.split('/')[-1].split('.')[0]
        sub_key = sub_url.split('/')[-1].split('.')[0]
        return img_key, sub_key
    except Exception as e:
        print(f"获取密钥异常，请检查网络或 Cookie: {e}")
        return None, None

def encWbi(params: dict, img_key: str, sub_key: str):
    """计算 Wbi 签名 (w_rid)"""
    mixin_key = getMixinKey(img_key + sub_key)
    curr_time = round(time.time())
    params['wts'] = curr_time
    # 按照 key 的字母顺序排序
    sorted_params = dict(sorted(params.items()))
    # 过滤掉非法字符并拼接成字符串
    query = urllib.parse.urlencode(sorted_params)
    # MD5 哈希计算 w_rid
    wbi_sign = hashlib.md5((query + mixin_key).encode()).hexdigest()
    sorted_params['w_rid'] = wbi_sign
    return sorted_params

# ================== 核心爬虫模块 ==================

def get_oid_from_bvid(bvid, headers):
    """通过 bvid 获取真实的 oid (aid)"""
    url = "https://api.bilibili.com/x/web-interface/view"
    params = {"bvid": bvid}
    try:
        resp = requests.get(url, headers=headers, params=params).json()
        if resp.get('code') == 0:
            return resp['data']['aid']
        else:
            print(f"[!] 获取 {bvid} 的 OID 失败: {resp.get('message')}")
            return None
    except Exception as e:
        print(f"[!] 请求视频详情异常 ({bvid}): {e}")
        return None

def fetch_sub_replies(root_id, oid, img_key, sub_key, headers):
    """
    抓取某条一级评论下的所有二级回复（楼中楼）
    返回一个包含多条回复信息的列表
    """
    sub_replies_data = []
    pn = 1
    
    while True:
        params = {
            "oid": oid,
            "type": 1,
            "root": root_id, # 一级评论的 rpid
            "ps": 20,
            "pn": pn,
            "web_location": 1315875
        }
        
        signed_params = encWbi(params, img_key, sub_key)
        api_url = "https://api.bilibili.com/x/v2/reply/reply"
        
        try:
            response = requests.get(api_url, headers=headers, params=signed_params, timeout=10)
            data = response.json()
            
            if data['code'] != 0:
                print(f"      [警告] 二级回复拉取异常 ({root_id}): {data['message']}")
                break
                
            replies = data['data'].get('replies')
            if not replies:
                break
                
            for reply in replies:
                uid = reply['member']['mid']
                uname = reply['member']['uname']
                message = reply['content']['message'].replace('\n', ' ')
                like_count = reply['like']
                rcount = reply['rcount'] # 楼中楼其实不太会再有回复，这里也可以只存0
                ctime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(reply['ctime']))
                
                # [层级标识, 用户ID, 用户名, 评论内容, 点赞数, 回复数, 发布时间, 父级ID]
                row = ["└ 回复", uid, uname, message, like_count, rcount, ctime, root_id]
                sub_replies_data.append(row)
                
            # 判断是否还要翻页
            page_info = data['data'].get('page', {})
            count = page_info.get('count', 0)  # 总评论数
            
            # 由于获取二级回复使用 pn 翻页，我们根据拉取的数据量和总数判断是否到底 (ps=20)
            if pn * 20 >= count:
                break
                
            pn += 1
            time.sleep(1) # 请求楼中楼的间隔，避免太快被封
            
        except Exception as e:
            print(f"      [错误] 抓取楼中楼失败 ({root_id}): {e}")
            break
            
    return sub_replies_data

def crawl_video_comments(bvid, config):
    """抓取单个视频的所有评论并导出 Excel"""
    print(f"\n{'='*50}")
    print(f"开始处理视频: {bvid}")
    
    headers = {
        "User-Agent": config.get("user_agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"),
        "Cookie": config.get("cookie", ""),
        "Referer": f"https://www.bilibili.com/video/{bvid}/"
    }
    
    oid = get_oid_from_bvid(bvid, headers)
    if not oid:
        print(f"跳过 {bvid}。")
        return
        
    print(f"成功获取 {bvid} 真实 OID (aid): {oid}")
    
    img_key, sub_key = get_wbi_keys(headers)
    if not img_key or not sub_key:
        print("停止抓取。")
        return

    # 全部评论行数据存储
    all_rows = []
    
    # === Excel 表头 ===
    # ['层级', '用户ID', '用户名', '评论内容', '点赞数', '回复数', '发布时间', '父评论ID']
    
    pagination_str = '{"offset":""}' 
    page_count = 1
    total_main = 0
    total_sub = 0

    while True:
        params = {
            "oid": oid,
            "type": 1,
            "mode": 2, # mode=2 按时间顺序最全 (mode=3热门)
            "pagination_str": pagination_str,
            "plat": 1,
            "web_location": 1315875
        }

        signed_params = encWbi(params, img_key, sub_key)
        api_url = "https://api.bilibili.com/x/v2/reply/main"
        
        try:
            response = requests.get(api_url, headers=headers, params=signed_params, timeout=10)
            data = response.json()

            if data['code'] != 0:
                print(f"[警告] 一级评论接口返回异常: {data['message']}")
                break

            replies = data['data'].get('replies')
            if not replies:
                print("[完成] 该视频已经没有更多一级评论可抓取了。")
                break

            print(f"[{bvid}] 正在抓取第 {page_count} 页一级评论 (本页 {len(replies)} 条) ...")
            
            for reply in replies:
                rpid = reply['rpid'] # 一级评论的唯一标识
                uid = reply['member']['mid']
                uname = reply['member']['uname']
                message = reply['content']['message'].replace('\n', ' ')
                like_count = reply['like']
                rcount = reply['rcount'] # 这个一级评论下的回复数
                ctime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(reply['ctime']))
                
                # 追加一级评论行
                main_row = ["▶ 评论", uid, uname, message, like_count, rcount, ctime, ""]
                all_rows.append(main_row)
                total_main += 1
                
                # 如果该评论下有子评论，去爬子评论
                if rcount > 0:
                    # 提示正在抓取楼中楼，防止长时间没输出以为卡死了
                    sub_rows = fetch_sub_replies(rpid, oid, img_key, sub_key, headers)
                    all_rows.extend(sub_rows)
                    total_sub += len(sub_rows)

            # 获取下一页标识
            cursor = data['data'].get('cursor', {})
            is_end = cursor.get('is_end')
            
            if is_end:
                print(f"\n[完成] 接口返回已经到底，{bvid} 评论抓取完毕！")
                break
                
            next_offset = cursor.get('pagination_reply', {}).get('next_offset')
            if not next_offset:
                break
                
            pagination_str = json.dumps({"offset": next_offset})
            page_count += 1
            time.sleep(2) # 休眠防止被封

        except KeyboardInterrupt:
            print("\n[中断] 用户手动终止了抓取，正在保存已获得的数据...")
            break
        except Exception as e:
            print(f"\n[错误] 请求一级评论页发生异常: {e}，等待5秒后继续...")
            time.sleep(5)

    print(f"\n[{bvid}] 数据收集完毕或中断！共 {total_main} 条一级评论，{total_sub} 条二级回复。")
    export_to_excel(bvid, all_rows)


def export_to_excel(bvid, rows):
    """将数据按照美观格式导出到 Excel 中"""
    if not rows:
        print(f"[{bvid}] 没有抓取到任何评论数据，跳过导出。")
        return
        
    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    os.makedirs(output_dir, exist_ok=True)
    
    excel_path = os.path.join(output_dir, f"bilibili_comments_{bvid}.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bvid
    
    headers = ['层级', '用户ID', '用户名', '评论内容', '点赞数', '回复数', '发布时间', '父评论ID']
    ws.append(headers)
    
    # 冻结第一行
    ws.freeze_panes = 'A2'
    
    # 表头样式: 加粗、背景灰色
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
    # 主评论样式: 淡蓝色背景
    main_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    
    for row_idx, row_data in enumerate(rows, start=2):
        ws.append(row_data)
        
        is_main = (row_data[0] == "▶ 评论")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # 一级评论整行涂蓝
            if is_main:
                cell.fill = main_fill
                
            # 二级回复的内容稍微缩进（加几个空格前缀显示）
            # 或者可以直接在这里调整缩进
            if not is_main and col_idx == 4: # 第4列是内容
                cell.value = "    " + str(cell.value)
                
            # 数字列靠右，其他靠左
            if col_idx in [2, 5, 6, 8]: # ID，点赞，回复数，父ID
                cell.alignment = Alignment(horizontal='right', vertical='top')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=(col_idx==4))

    # 调整列宽
    ws.column_dimensions['A'].width = 10 # 层级
    ws.column_dimensions['B'].width = 15 # 用户ID
    ws.column_dimensions['C'].width = 20 # 用户名
    ws.column_dimensions['D'].width = 80 # 评论内容
    ws.column_dimensions['E'].width = 8  # 点赞
    ws.column_dimensions['F'].width = 8  # 回复
    ws.column_dimensions['G'].width = 20 # 时间
    ws.column_dimensions['H'].width = 15 # 父评论ID

    wb.save(excel_path)
    print(f"[{bvid}] Excel 导出成功: {excel_path}")


def main():
    print("======== B站评论爬虫工具 (支持两级评论 & 批量) ========")
    
    config = load_config()
    if not config:
        return
        
    bvid_list = load_bvid_list()
    if not bvid_list:
        return
        
    print(f"检查到待爬取视频数量: {len(bvid_list)}")
    
    for idx, bvid in enumerate(bvid_list, 1):
        print(f"\n[任务 {idx}/{len(bvid_list)}] 准备处理: {bvid}")
        crawl_video_comments(bvid, config)
        
    print("\n🎉 所有任务已处理完成！")

if __name__ == "__main__":
    main()
